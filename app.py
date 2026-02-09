"""
Gold Portfolio Tracker - Backend API
Fetches gold prices from Galeri24.co.id and manages portfolio
"""

from flask import Flask, jsonify, request, send_from_directory, Response
from flask_cors import CORS
import requests
from bs4 import BeautifulSoup
from decimal import Decimal
from datetime import datetime
from zoneinfo import ZoneInfo
import json
import os
import csv
import io
import database as db
from apscheduler.schedulers.background import BackgroundScheduler
import atexit

app = Flask(__name__, static_folder='static')
CORS(app)

# Initialize background scheduler for hourly price tracking
scheduler = BackgroundScheduler(timezone="Asia/Jakarta")

def record_hourly_price():
    """Background job: Fetch and record 1 gram gold price if changed"""
    try:
        prices = get_gold_prices()
        if prices and prices.get("success"):
            data = prices.get("data", {})
            # Get 1 gram price (try both "1.0" and "1" keys)
            one_gram = data.get("1.0") or data.get("1")
            if one_gram:
                changed = db.save_price_history(
                    weight=1.0,
                    sell_price=one_gram["sell"],
                    buy_price=one_gram["buy"]
                )
                if changed:
                    print(f"✅ Price updated at {datetime.now(ZoneInfo('Asia/Jakarta')).strftime('%Y-%m-%d %H:%M:%S')}: Sell={one_gram['sell']}, Buy={one_gram['buy']}")
                else:
                    print(f"ℹ️  Price unchanged at {datetime.now(ZoneInfo('Asia/Jakarta')).strftime('%Y-%m-%d %H:%M:%S')}")
    except Exception as e:
        print(f"❌ Price recording error: {e}")

# Schedule hourly price check (at minute 0 of every hour: 9:00, 10:00, 11:00, etc.)
scheduler.add_job(
    func=record_hourly_price,
    trigger="cron",
    minute=0,  # Run at xx:00
    id="hourly_price_check",
    replace_existing=True
)

# Start scheduler
scheduler.start()

# Shutdown scheduler gracefully on app exit
atexit.register(lambda: scheduler.shutdown())

# Record price immediately on startup
record_hourly_price()


def clean_price(price_str):
    """Clean price string like 'Rp1.041.000' -> Decimal('1041000')"""
    price_str = price_str.replace("Rp", "").replace(".", "").replace(",", "").strip()
    return Decimal(price_str)

def decimal_to_float(obj):
    """Convert Decimal to float for JSON serialization"""
    if isinstance(obj, Decimal):
        return float(obj)
    raise TypeError

def get_gold_prices():
    """Fetch current gold prices from Galeri24.co.id"""
    url = "https://galeri24.co.id/harga-emas"
    
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        }
        response = requests.get(url, headers=headers, timeout=15)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.text, "html.parser")
        galeri24_div = soup.find("div", {"id": "GALERI 24"})
        
        if not galeri24_div:
            return None
            
        main_container = galeri24_div.find("div", class_="grid divide-neutral-200 border-neutral-200")
        
        if not main_container:
            return None
            
        rows = main_container.find_all("div", class_="grid grid-cols-5 divide-x lg:hover:bg-neutral-50 transition-all")
        
        gold_prices = {}
        
        for row in rows:
            cols = row.find_all("div", class_="p-3 col-span-1 whitespace-nowrap w-fit") + \
                   row.find_all("div", class_="p-3 col-span-2 whitespace-nowrap w-fit")
            cols_text = [col.get_text(strip=True) for col in cols]
            
            if len(cols_text) == 3:
                weight = float(cols_text[0])
                sell_price = float(clean_price(cols_text[1]))
                buy_price = float(clean_price(cols_text[2]))
                spread_pct = ((sell_price - buy_price) / buy_price * 100) if buy_price else 0
                
                gold_prices[str(weight)] = {
                    "weight": weight,
                    "sell": sell_price,
                    "buy": buy_price,
                    "spread_pct": round(spread_pct, 2)
                }
        
        tz = ZoneInfo("Asia/Jakarta")
        last_update = datetime.now(tz).strftime('%Y-%m-%d %H:%M:%S')
        
        return {
            "success": True,
            "last_update": last_update,
            "timezone": "Asia/Jakarta (GMT+7)",
            "data": gold_prices
        }
        
    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }

def load_portfolio():
    """Load portfolio from database"""
    return db.load_portfolio()

def save_portfolio(portfolio):
    """Save portfolio - now handled by database module"""
    # This is now handled by individual save operations
    pass

# ============== API ROUTES ==============

@app.route('/')
def serve_index():
    return send_from_directory('static', 'index.html')

@app.route('/static/<path:path>')
def serve_static(path):
    return send_from_directory('static', path)

@app.route('/api/prices', methods=['GET'])
def api_get_prices():
    """Get current gold prices"""
    prices = get_gold_prices()
    return jsonify(prices)

@app.route('/api/portfolio', methods=['GET'])
def api_get_portfolio():
    """Get user's portfolio"""
    portfolio = load_portfolio()
    return jsonify({"success": True, "data": portfolio})

@app.route('/api/portfolio/holdings', methods=['POST'])
def api_add_holding():
    """Add a new gold holding"""
    data = request.json
    
    holding = {
        "id": datetime.now().strftime('%Y%m%d%H%M%S%f'),
        "weight": float(data.get('weight', 0)),
        "purchase_price": float(data.get('purchase_price', 0)),
        "purchase_date": data.get('purchase_date', datetime.now().strftime('%Y-%m-%d')),
        "notes": data.get('notes', ''),
        "created_at": datetime.now(ZoneInfo("Asia/Jakarta")).isoformat()
    }
    
    # Save to database
    db.save_holding(holding)
    
    # Add transaction record
    db.save_transaction({
        "type": "BUY",
        "holding_id": holding["id"],
        "weight": holding["weight"],
        "price": holding["purchase_price"],
        "date": holding["purchase_date"],
        "timestamp": holding["created_at"]
    })
    
    return jsonify({"success": True, "data": holding})

@app.route('/api/portfolio/holdings/<holding_id>', methods=['DELETE'])
def api_delete_holding(holding_id):
    """Delete a gold holding (sell or delete)"""
    data = request.json or {}
    sell_price = data.get('sell_price', 0)
    
    # Delete from database
    deleted = db.delete_holding(holding_id, record_transaction=True, sell_price=sell_price)
    
    if deleted:
        msg = "Holding sold successfully" if sell_price > 0 else "Holding deleted successfully"
        return jsonify({"success": True, "message": msg})
    
    return jsonify({"success": False, "error": "Holding not found"}), 404

@app.route('/api/portfolio/holdings/<holding_id>', methods=['PUT'])
def api_update_holding(holding_id):
    """Update a gold holding"""
    data = request.json
    
    updates = {
        "weight": float(data.get('weight', 0)),
        "purchase_price": float(data.get('purchase_price', 0)),
        "purchase_date": data.get('purchase_date', ''),
        "notes": data.get('notes', '')
    }
    # Remove empty values
    updates = {k: v for k, v in updates.items() if v}
    
    updated = db.update_holding(holding_id, updates)
    
    if updated:
        return jsonify({"success": True, "data": updated})
    
    return jsonify({"success": False, "error": "Holding not found"}), 404

@app.route('/api/portfolio/export', methods=['GET'])
def api_export_portfolio():
    """Export portfolio to CSV"""
    csv_data = db.export_to_csv()
    return Response(
        csv_data,
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment;filename=gold_portfolio.csv'}
    )

@app.route('/api/portfolio/import', methods=['POST'])
def api_import_holdings():
    """Import holdings from CSV or Excel file"""
    if 'file' not in request.files:
        return jsonify({"success": False, "error": "No file uploaded"}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({"success": False, "error": "No file selected"}), 400
    
    filename = file.filename.lower()
    imported_count = 0
    errors = []
    
    try:
        if filename.endswith('.csv'):
            # Parse CSV
            content = file.read().decode('utf-8')
            reader = csv.DictReader(io.StringIO(content))
            rows = list(reader)
        elif filename.endswith(('.xlsx', '.xls')):
            # Parse Excel
            try:
                from openpyxl import load_workbook
                wb = load_workbook(file)
                ws = wb.active
                headers = [cell.value.lower().strip() if cell.value else '' for cell in ws[1]]
                rows = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
                    rows.append(row_dict)
            except ImportError:
                return jsonify({"success": False, "error": "Excel support not installed"}), 400
        else:
            return jsonify({"success": False, "error": "Unsupported file format. Use CSV or Excel (.xlsx)"}), 400
        
        # Column name mappings (flexible matching)
        weight_cols = ['weight', 'berat', 'gram', 'gr']
        price_cols = ['purchase price', 'purchase_price', 'price', 'harga', 'harga_beli', 'cost', 'total purchase cost', 'total_purchase_cost']
        date_cols = ['purchase date', 'purchase_date', 'date', 'tanggal', 'tanggal_beli']
        notes_cols = ['notes', 'note', 'catatan', 'keterangan', 'type', 'jenis']
        quantity_cols = ['quantity', 'qty', 'jumlah']
        
        def find_value(row, col_names):
            row_lower = {k.lower().strip(): v for k, v in row.items() if k}
            for col in col_names:
                if col in row_lower and row_lower[col] is not None:
                    return row_lower[col]
            return None
        
        for i, row in enumerate(rows):
            try:
                weight = find_value(row, weight_cols)
                price = find_value(row, price_cols)
                date = find_value(row, date_cols)
                notes = find_value(row, notes_cols) or ''
                quantity = find_value(row, quantity_cols)
                
                if not weight or not price:
                    errors.append(f"Row {i+2}: Missing weight or price")
                    continue
                
                # Clean weight (remove 'g', 'gr', 'gram')
                weight_str = str(weight).lower().replace('g', '').replace('r', '').replace('a', '').replace('m', '').strip()
                weight_val = float(weight_str)
                
                # Clean price (remove 'Rp', '.', ',', quotes, spaces)
                price_str = str(price).replace('Rp', '').replace('"', '').replace("'", '').strip()
                # Handle both formats: "Rp42,118,275" and "42.118.275"
                if ',' in price_str and '.' not in price_str:
                    # Comma as thousands separator: "42,118,275"
                    price_str = price_str.replace(',', '')
                elif '.' in price_str and ',' not in price_str:
                    # Dot as thousands separator: "42.118.275"
                    price_str = price_str.replace('.', '')
                else:
                    # Both present, assume comma is thousands: "42,118.00" or just clean all
                    price_str = price_str.replace(',', '').replace('.', '')
                price_val = float(price_str) if price_str else 0
                
                # Parse quantity (default 1)
                quantity_val = 1
                if quantity:
                    try:
                        quantity_val = int(float(str(quantity)))
                    except:
                        quantity_val = 1
                
                # Parse date
                if date:
                    if hasattr(date, 'strftime'):
                        date_str = date.strftime('%Y-%m-%d')
                    else:
                        date_str = str(date)[:10]
                else:
                    date_str = datetime.now(ZoneInfo("Asia/Jakarta")).strftime('%Y-%m-%d')
                
                # Create holdings for each quantity
                for q in range(quantity_val):
                    holding = {
                        "id": datetime.now().strftime('%Y%m%d%H%M%S%f') + str(i) + str(q),
                        "weight": weight_val,
                        "purchase_price": price_val,
                        "purchase_date": date_str,
                        "notes": str(notes) if notes else f"{weight_val}g",
                        "created_at": datetime.now(ZoneInfo("Asia/Jakarta")).isoformat()
                    }
                    
                    # Save to database
                    db.save_holding(holding)
                    db.save_transaction({
                        "type": "BUY",
                        "holding_id": holding["id"],
                        "weight": holding["weight"],
                        "price": holding["purchase_price"],
                        "date": holding["purchase_date"],
                        "timestamp": holding["created_at"]
                    })
                    imported_count += 1
                    
            except Exception as e:
                errors.append(f"Row {i+2}: {str(e)}")
        
        return jsonify({
            "success": True,
            "imported": imported_count,
            "errors": errors[:10]  # Limit errors shown
        })
        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/portfolio/summary', methods=['GET'])
def api_portfolio_summary():
    """Get portfolio summary with current valuations"""
    portfolio = load_portfolio()
    prices = get_gold_prices()
    
    if not prices.get("success"):
        return jsonify({"success": False, "error": "Could not fetch current prices"})
    
    total_weight = 0
    total_cost = 0
    total_current_value = 0
    holdings_with_values = []
    
    for holding in portfolio["holdings"]:
        weight = holding["weight"]
        cost = holding["purchase_price"]
        
        # Find matching price or estimate per gram
        price_key = str(weight)
        if price_key in prices["data"]:
            current_sell = prices["data"][price_key]["sell"]
            current_buy = prices["data"][price_key]["buy"]
        else:
            # Estimate using 1 gram price
            per_gram = prices["data"].get("1.0", prices["data"].get("1", {}))
            if per_gram:
                current_sell = per_gram["sell"] * weight
                current_buy = per_gram["buy"] * weight
            else:
                current_sell = 0
                current_buy = 0
        
        profit_loss = current_buy - cost
        profit_loss_pct = ((current_buy - cost) / cost * 100) if cost > 0 else 0
        
        holdings_with_values.append({
            **holding,
            "current_sell": current_sell,
            "current_buy": current_buy,
            "profit_loss": profit_loss,
            "profit_loss_pct": round(profit_loss_pct, 2)
        })
        
        total_weight += weight
        total_cost += cost
        total_current_value += current_buy
    
    total_profit_loss = total_current_value - total_cost
    total_profit_loss_pct = ((total_current_value - total_cost) / total_cost * 100) if total_cost > 0 else 0
    
    return jsonify({
        "success": True,
        "prices_update": prices["last_update"],
        "summary": {
            "total_weight": round(total_weight, 2),
            "total_cost": round(total_cost, 0),
            "total_current_value": round(total_current_value, 0),
            "total_profit_loss": round(total_profit_loss, 0),
            "total_profit_loss_pct": round(total_profit_loss_pct, 2),
            "holdings_count": len(portfolio["holdings"])
        },
        "holdings": holdings_with_values,
        "transactions": portfolio["transactions"]
    })

@app.route('/api/price-history', methods=['GET'])
def api_price_history():
    """Get price history for 1 gram gold"""
    days = request.args.get('days', default=30, type=int)
    # Limit to reasonable range
    days = max(1, min(days, 365))
    
    history = db.get_price_history(weight=1.0, days=days)
    
    return jsonify({
        "success": True,
        "days": days,
        "count": len(history),
        "data": history
    })

if __name__ == '__main__':
    # Create static folder if not exists
    os.makedirs('static', exist_ok=True)
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_DEBUG', 'True').lower() == 'true'
    app.run(debug=debug, host='0.0.0.0', port=port)
